"""
Feeder (bus-to-bus) data integration — visualization only.
Reads engineering feeder data from Excel (Bus IDs), maps to posts via BusPostMapping,
returns line segments to draw only when both buses map to posts with coordinates.
Does not infer or apply any electrical logic.
"""
import os
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def _normalize_bus_id(value):
    """Normalize bus id for lookup (string, stripped)."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def get_bus_to_post_map(app):
    """
    Build dict bus_id -> post_id from BusPostMapping table.
    Uses normalized string keys for bus_id.
    Returns {} if table does not exist or query fails (e.g. migration not run).
    """
    try:
        from models import BusPostMapping
        with app.app_context():
            rows = BusPostMapping.query.all()
            return {_normalize_bus_id(r.bus_id): r.post_id for r in rows if _normalize_bus_id(r.bus_id) is not None}
    except Exception as e:
        logger.warning("Bus–Post mapping unavailable (table may not exist): %s", e)
        return {}


def get_post_coords_by_id(app):
    """Build dict post_id -> (lat, lng) from Post table. Only posts with valid coords."""
    try:
        from models import Post
        with app.app_context():
            posts = Post.query.all()
            out = {}
            for p in posts:
                try:
                    lat, lng = float(p.lat), float(p.lng)
                    out[p.id] = (lat, lng)
                except (TypeError, ValueError):
                    continue
            return out
    except Exception as e:
        logger.warning("Post coords unavailable: %s", e)
        return {}


def read_feeder_excel(path):
    """
    Read bus-to-bus connections from an Excel file.
    Expects columns for from_bus and to_bus (case-insensitive): e.g. From_Bus, To_Bus.
    Returns list of (from_bus, to_bus) tuples (values normalized as strings).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; feeder Excel cannot be read")
        return []

    if not path or not os.path.isfile(path):
        logger.info("Feeder Excel not found or not a file: %s", path)
        return []

    pairs = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        # Use first sheet
        ws = wb.active
        if ws is None:
            wb.close()
            return []

        # First row: find column indices for from_bus and to_bus (case-insensitive)
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        lower_headers = [str(h).strip().lower() if h is not None else "" for h in header]
        from_keys = ("from_bus", "from bus", "frombus", "bus_from", "bus from")
        to_keys = ("to_bus", "to bus", "tobus", "bus_to", "bus to")
        from_col = None
        to_col = None
        for i, h in enumerate(lower_headers):
            if h in from_keys:
                from_col = i
                break
        for i, h in enumerate(lower_headers):
            if h in to_keys:
                to_col = i
                break
        # Fallback: first two columns
        if from_col is None:
            from_col = 0
        if to_col is None:
            to_col = 1

        for row in ws.iter_rows(min_row=2):
            vals = [cell.value for cell in row]
            from_bus = _normalize_bus_id(vals[from_col]) if from_col < len(vals) else None
            to_bus = _normalize_bus_id(vals[to_col]) if to_col < len(vals) else None
            if from_bus is None and to_bus is None:
                continue
            if from_bus == to_bus:
                continue
            pairs.append((from_bus or "", to_bus or ""))
        wb.close()
    except Exception as e:
        logger.exception("Error reading feeder Excel %s: %s", path, e)
        return []

    return pairs


def resolve_feeder_lines(app):
    """
    Resolve engineering feeder (bus-to-bus) data to drawable lines using Bus–Post mapping.
    - Reads bus–post mapping from DB and post coordinates from Post table.
    - Reads bus-to-bus pairs from Excel (path from FEEDER_EXCEL_PATH or data/feeder_connections.xlsx).
    - Only includes a line when both buses map to a post that has coordinates.
    - Logs and returns skipped connections (unmapped bus or missing coords).
    Returns: {"lines": [...], "skipped": [...]}
    """
    bus_to_post = get_bus_to_post_map(app)
    post_coords = get_post_coords_by_id(app)

    # Path to feeder Excel: env or default under project data/
    base_dir = Path(app.root_path)
    default_path = base_dir / "data" / "feeder_connections.xlsx"
    excel_path = os.getenv("FEEDER_EXCEL_PATH") or str(default_path)
    pairs = read_feeder_excel(excel_path)

    lines = []
    skipped = []

    for from_bus, to_bus in pairs:
        if not from_bus or not to_bus:
            reason = "empty bus id"
            skipped.append({"from_bus": from_bus, "to_bus": to_bus, "reason": reason})
            logger.debug("Feeder skipped %s -> %s: %s", from_bus, to_bus, reason)
            continue

        post_id_from = bus_to_post.get(from_bus)
        post_id_to = bus_to_post.get(to_bus)

        if post_id_from is None:
            reason = "from_bus not in bus–post mapping"
            skipped.append({"from_bus": from_bus, "to_bus": to_bus, "reason": reason})
            logger.info("Feeder skipped %s -> %s: %s", from_bus, to_bus, reason)
            continue
        if post_id_to is None:
            reason = "to_bus not in bus–post mapping"
            skipped.append({"from_bus": from_bus, "to_bus": to_bus, "reason": reason})
            logger.info("Feeder skipped %s -> %s: %s", from_bus, to_bus, reason)
            continue

        coords_from = post_coords.get(post_id_from)
        coords_to = post_coords.get(post_id_to)
        if coords_from is None:
            reason = "post for from_bus has no coordinates"
            skipped.append({"from_bus": from_bus, "to_bus": to_bus, "reason": reason})
            logger.info("Feeder skipped %s -> %s: %s", from_bus, to_bus, reason)
            continue
        if coords_to is None:
            reason = "post for to_bus has no coordinates"
            skipped.append({"from_bus": from_bus, "to_bus": to_bus, "reason": reason})
            logger.info("Feeder skipped %s -> %s: %s", from_bus, to_bus, reason)
            continue

        lat1, lng1 = coords_from
        lat2, lng2 = coords_to
        lines.append({
            "from_bus": from_bus,
            "to_bus": to_bus,
            "post_id_from": post_id_from,
            "post_id_to": post_id_to,
            "lat1": lat1, "lng1": lng1,
            "lat2": lat2, "lng2": lng2,
        })

    return {"lines": lines, "skipped": skipped}
