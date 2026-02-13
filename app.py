from flask import Flask, render_template, jsonify, request, redirect, url_for, session, g, Response
from flask_cors import CORS
from dotenv import load_dotenv
import os
import traceback
import secrets
import csv
import io
import re
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

from extensions import db, migrate

load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')
# Secret key: use configured SECRET_KEY or generate a temporary one (not recommended for production)
app.secret_key = os.getenv('SECRET_KEY') or os.urandom(24)
CORS(app)

# Dev UX: auto-reload templates and disable static caching when configured
if str(os.getenv('TEMPLATES_AUTO_RELOAD', '')).lower() in ('1', 'true', 'yes'):
    app.config['TEMPLATES_AUTO_RELOAD'] = True
    app.jinja_env.auto_reload = True
if str(os.getenv('SEND_FILE_MAX_AGE_DEFAULT', '')).strip() != '':
    try:
        app.config['SEND_FILE_MAX_AGE_DEFAULT'] = int(os.getenv('SEND_FILE_MAX_AGE_DEFAULT', '0'))
    except Exception:
        app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Configure database (MySQL recommended for production)
# Support either a single DATABASE_URL or individual DB_* environment variables
database_url = os.getenv('DATABASE_URL')
if not database_url:
    db_user = os.getenv('DB_USERNAME')
    db_pass = os.getenv('DB_PASSWORD')
    db_host = os.getenv('DB_HOST', '127.0.0.1')
    db_port = os.getenv('DB_PORT', '3306')
    db_name = os.getenv('DB_DATABASE')
    if db_user and db_name is not None:
        # Build a sqlalchemy URL for PyMySQL driver
        if db_pass:
            database_url = f"mysql+pymysql://{db_user}:{db_pass}@{db_host}:{db_port}/{db_name}"
        else:
            database_url = f"mysql+pymysql://{db_user}@{db_host}:{db_port}/{db_name}"
    else:
        database_url = 'sqlite:///app.db'

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Bind extensions to app
db.init_app(app)
migrate.init_app(app, db)

# Import models after DB is available to avoid circular imports
with app.app_context():
    import models  # noqa: E402,F401
    # Bring commonly used model names into this module namespace
    from models import User, Post, Meter, BusPostMapping, DistributionLineSegment, LineConnection, DistributionTransformer  # noqa: E402,F401

# Helper: get current user from session
def get_current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    try:
        return User.query.get(uid)
    except Exception:
        return None

# Make `current_user` available in templates
@app.context_processor
def inject_current_user():
    return {'current_user': get_current_user()}

@app.before_request
def load_current_user_into_g():
    g.current_user = get_current_user()

# Role/permission decorators
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = g.get('current_user')
        if not u or not u.is_admin():
            # API calls: return JSON; otherwise redirect to login
            if request.path.startswith('/api/'):
                return jsonify({'error': 'admin required'}), 403
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return wrapper

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not g.get('current_user'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'authentication required'}), 401
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return wrapper
# Sample in-memory posts — used as a fallback if DB is empty
POSTS = [
    {"id": 1, "name": "Pole A", "lat": 14.5995, "lng": 120.9842, "status": "Active"},
    {"id": 2, "name": "Pole B", "lat": 14.6091, "lng": 121.0223, "status": "Active"},
]

@app.route('/')
def index():
    """Root path.

    If authenticated, send users to the map page.
    Otherwise send users to the login page.
    """
    u = get_current_user()
    if u:
        return redirect(url_for('map_view'))
    return redirect(url_for('login'))


@app.route('/map')
@login_required
def map_view():
    """Interactive map (protected route)."""
    return render_template('index.html', active_page='map')


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard (protected route).

    This route is protected so that unauthenticated users cannot view or interact with the
    dashboard UI. The `@login_required` decorator enforces authentication on the server
    side — if an unauthenticated user tries to access /dashboard via direct URL, they are
    redirected to /login. This prevents relying on client-side hiding of elements for
    security (client UI is only a convenience); the backend is authoritative for access.
    """
    try:
        total = Post.query.count()
        in_ph = Post.query.filter(Post.lat >= 4.0, Post.lat <= 22.0, Post.lng >= 116.0, Post.lng <= 127.5).count()
        recent = Post.query.order_by(Post.id.desc()).all()  # Fetch all posts, not just 10
    except Exception:
        total = 0
        in_ph = 0
        recent = []
    return render_template('dashboard.html', active_page='dashboard', stats={'total': total, 'in_ph': in_ph}, posts=recent)

@app.route('/seed', methods=['POST'])
def seed():
    from seed_db import seed_posts
    seed_posts()
    return redirect(url_for('dashboard'))

@app.route('/api/posts', methods=['GET', 'POST'])
def api_posts():
    """GET: return posts. Optional query arg in_ph=1 filters to Philippines bbox.
    POST: create a new post. Body: { name: str, lat: float, lng: float, status?: str, area?: str }
    """
    if request.method == 'POST':
        # Only admins can create posts
        if not g.get('current_user') or not g.current_user.is_admin():
            return jsonify({'error': 'admin required'}), 403
        try:
            data = request.get_json() or {}
            name = (data.get('name') or '').strip()
            if not name:
                return jsonify({'error': 'name is required'}), 400
            try:
                lat = float(data.get('lat'))
                lng = float(data.get('lng'))
            except (TypeError, ValueError):
                return jsonify({'error': 'lat and lng must be numbers'}), 400
            if not (-90 <= lat <= 90 and -180 <= lng <= 180):
                return jsonify({'error': 'lat/lng out of range'}), 400
            status = (data.get('status') or '').strip() or None
            area = (data.get('area') or '').strip() or None
            post = Post(name=name, lat=lat, lng=lng, status=status, area=area)
            db.session.add(post)
            db.session.commit()
            return jsonify({
                'id': post.id,
                'name': post.name,
                'lat': post.lat,
                'lng': post.lng,
                'status': post.status,
            }), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # GET
    in_ph = str(request.args.get('in_ph', '')).lower() in ('1', 'true', 'yes')
    
    # Pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # Validate pagination parameters
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:  # Cap per_page at 1000 to support full map requests
        per_page = 10
    
    try:
        query = Post.query
        if in_ph:
            # Filter in DB for performance
            query = query.filter(Post.lat >= 4.0, Post.lat <= 22.0, Post.lng >= 116.0, Post.lng <= 127.5)
        
        # Get total count before pagination
        total = query.count()
        
        # Apply pagination
        db_posts = query.offset((page - 1) * per_page).limit(per_page).all()
        posts = [{"id": p.id, "name": p.name, "lat": p.lat, "lng": p.lng, "status": p.status} for p in db_posts]
        
        # Calculate pagination metadata
        total_pages = (total + per_page - 1) // per_page  # Ceiling division
        
        # Return paginated response
        response = {
            "data": posts,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            }
        }
        
        # Important: if DB query succeeds, return DB result even if empty.
        # This prevents silently falling back to demo data when the DB has no rows.
        return jsonify(response)
    except Exception as e:
        # In dev, surface the error so you know why DB isn't being used.
        if app.debug:
            return jsonify({
                "error": "DB query failed",
                "details": str(e),
                "db_uri": app.config.get("SQLALCHEMY_DATABASE_URI"),
                "trace": traceback.format_exc(),
            }), 500
        # In prod, fall back silently
        pass
    
    # Fallback to in-memory POSTS (optionally filtered) only on exception
    fallback_posts = POSTS
    if in_ph:
        fallback_posts = [p for p in POSTS if p['lat'] >= 4.0 and p['lat'] <= 22.0 and p['lng'] >= 116.0 and p['lng'] <= 127.5]
    
    total = len(fallback_posts)
    total_pages = (total + per_page - 1) // per_page
    paginated_posts = fallback_posts[(page - 1) * per_page:page * per_page]
    
    return jsonify({
        "data": paginated_posts,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        }
    })


@app.route('/api/posts/<int:post_id>', methods=['GET', 'PUT', 'DELETE'])
def api_post_detail(post_id):
    """GET: Return a single post by ID.
    PUT: Update post (admin only). Body: { lat?: float, lng?: float, status?: str }
    DELETE: Delete post (admin only).
    """
    try:
        p = Post.query.get(post_id)
        if not p:
            return jsonify({'error': 'post not found'}), 404
        
        if request.method == 'GET':
                # Return full post data from database (authoritative `post` table)
                post_dict = p.to_dict()
                # Include meter history
                meters = []
                for m in getattr(p, 'meters', []):
                    meters.append({
                        'id': m.id,
                        'meter_id': m.meter_id,
                        'meter_brand': m.meter_brand,
                        'meter_rating': m.meter_rating,
                        'kwhr_reading': m.kwhr_reading,
                        'reading_date': m.reading_date.isoformat() if m.reading_date else None,
                    })
                post_dict['meters'] = meters
                # Add remaining infrastructure fields not included by to_dict()
                extra_fields = ['pri_structure','pri_conductor_size','neutral_wire','configuration',
                                'primary_bus_id','sec_structure','sec_conductor_size','sec_type',
                                'conductor_type','sec_bus_id','common_sole','transformer_bus_id',
                                'transformer_phasing','grounding_rod','l2_wire_type','l1_wire_type',
                                'created_at','updated_at']
                for f in extra_fields:
                    post_dict[f] = getattr(p, f, None)
                return jsonify(post_dict)
        
        # PUT and DELETE require admin
        if not g.get('current_user') or not g.current_user.is_admin():
            return jsonify({'error': 'admin required'}), 403
        
        if request.method == 'PUT':
            data = request.get_json() or {}
            if 'lat' in data:
                p.lat = float(data['lat'])
            if 'lng' in data:
                p.lng = float(data['lng'])
            if 'status' in data:
                p.status = (data['status'] or '').strip() or None
            db.session.commit()
            return jsonify({'success': True, 'id': p.id, 'name': p.name, 'lat': p.lat, 'lng': p.lng, 'status': p.status})
        
        elif request.method == 'DELETE':
            db.session.delete(p)
            db.session.commit()
            return jsonify({'success': True})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/posts/<int:post_id>/connections', methods=['GET'])
def api_post_connections(post_id):
    """Return inferred LineConnection records related to this post.
    Matches connections where the post's primary/secondary/transformer bus IDs
    (or generated P{pole_number}) appear as either from_bus or to_bus.
    Returns an empty array if the post is not found or no connections exist.
    """
    try:
        p = Post.query.get(post_id)
        if not p:
            return jsonify([]), 200

        buses = set()
        if p.primary_bus_id:
            buses.add(p.primary_bus_id)
        if p.sec_bus_id:
            buses.add(p.sec_bus_id)
        if p.transformer_bus_id:
            buses.add(p.transformer_bus_id)
        # Include generated bus id fallback used elsewhere in the app
        try:
            buses.add(f"P{p.pole_number}")
        except Exception:
            pass

        if not buses:
            return jsonify([]), 200

        # Query connections where either endpoint matches any of the post's buses
        conns = LineConnection.query.filter(
            (LineConnection.from_bus.in_(list(buses))) | (LineConnection.to_bus.in_(list(buses)))
        ).order_by(LineConnection.id.asc()).all()

        return jsonify([c.to_dict() for c in conns])
    except Exception as e:
        app.logger.warning('Failed to fetch connections for post %s: %s', post_id, e)
        return jsonify([]), 200


def normalize_column_name(name):
    """Normalize column name for flexible matching.

    - Lowercase
    - Collapse any non-alphanumeric (spaces, punctuation, parentheses) to single underscores
    - Trim leading/trailing underscores

    This makes headers like 'Length         (meters)' normalize to 'length_meters',
    which can match our DB fields like 'length_meters'.
    """
    if not name:
        return ""
    text = str(name).strip().lower()
    # Replace any run of non-alphanumeric chars with a single underscore
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return text.strip('_')

def find_column_value(row, possible_names, header_map=None):
    """
    Find column value using flexible matching of column names.
    Works with both dict rows (CSV) and tuple rows (Excel).
    """
    if not row or not possible_names:
        return None
    
    # For CSV (dict) rows
    if isinstance(row, dict):
        normalized_targets = [normalize_column_name(n) for n in possible_names]
        for key in row.keys():
            normalized_key = normalize_column_name(key)
            if normalized_key in normalized_targets:
                val = row[key]
                return str(val).strip() if val else None
    
    # For Excel (tuple) rows with header_map
    elif header_map:
        for name in possible_names:
            normalized_name = normalize_column_name(name)
            for header_key, idx in header_map.items():
                if normalize_column_name(header_key) == normalized_name:
                    if 0 <= idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val else None
    
    return None

def sanitize_float(value):
    """Convert value to float, return None if invalid"""
    if not value or str(value).strip() == '':
        return None
    try:
        return float(value)
    except (ValueError, AttributeError, TypeError):
        return None

@app.route('/api/posts/bulk-import', methods=['POST'])
@admin_required
def api_posts_bulk_import():
    """
    Bulk import posts from CSV or Excel file with flexible column name handling.
    Supports infrastructure data: pole_number, feeder, transformers, meters, etc.
    
    Required columns: Pole Number (or pole_number), Latitude/Lat, Longitude/Long
    Optional: All infrastructure columns (Feeder, kVA Rating, Meter Brand, etc.)
    
    Returns: { created: int, updated: int, skipped: int, errors: [{row, error}] }
    """
    from werkzeug.utils import secure_filename
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    filename = secure_filename(file.filename).lower()
    is_excel = filename.endswith(('.xlsx', '.xls'))
    is_csv = filename.endswith('.csv')
    
    if not (is_excel or is_csv):
        return jsonify({'error': 'File must be CSV or Excel (.xlsx, .xls)'}), 400
    
    try:
        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': [], 'message': 'Import successful'}
        rows = []
        header_map = {}

        # Parse file
        if is_excel:
            from openpyxl import load_workbook
            workbook = load_workbook(file.stream)
            worksheet = workbook.active
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    for idx, col_name in enumerate(row):
                        if col_name:
                            header_map[str(col_name).strip()] = idx
                    continue
                rows.append(row)
        else:  # CSV
            import csv as csv_module
            stream = io.StringIO(file.stream.read().decode('utf-8', errors='replace'))
            reader = csv_module.DictReader(stream)
            if not reader.fieldnames:
                return jsonify({'error': 'CSV file is empty or invalid'}), 400
            rows = list(reader)

        if not rows:
            return jsonify({'error': 'File has no data rows'}), 400

        # --- Unified Import: Map all fields to correct tables ---
        from models import Post, DistributionLineSegment
        from sqlalchemy.exc import IntegrityError
        post_fields = set([c.name for c in Post.__table__.columns])
        line_fields = set([c.name for c in DistributionLineSegment.__table__.columns])
        seen_posts = set()
        seen_lines = set()
        for row_idx, row in enumerate(rows, start=2):
            try:
                # Flexible: support both dict (CSV) and tuple (Excel)
                def get_val(possible, default=None):
                    return find_column_value(row, possible, header_map) or default

                # --- POST FIELDS ---
                post_data = {}
                for field in post_fields:
                    val = get_val([field, field.replace('_', ' '), field.upper(), field.lower()])
                    if val is not None:
                        post_data[field] = val

                # Extra explicit mappings for engineering fields whose headers
                # in EXAMPLEDATA.csv include units in parentheses.
                # These only run if the Post model actually has the field.
                if 'length_meters' in post_fields and 'length_meters' not in post_data:
                    post_data['length_meters'] = sanitize_float(
                        get_val(['Length (meters)', 'Length         (meters)'])
                    )

                if 'conductor_unit' in post_fields and 'conductor_unit' not in post_data:
                    post_data['conductor_unit'] = get_val(['Unit (C)', 'Unit (C) '])

                if 'conductor_strands' in post_fields and 'conductor_strands' not in post_data:
                    post_data['conductor_strands'] = get_val(['Strands (C)', 'Strands (C) '])

                if 'neutral_wire_unit' in post_fields and 'neutral_wire_unit' not in post_data:
                    post_data['neutral_wire_unit'] = get_val(['Unit (NW)', 'Unit (NW) '])

                if 'neutral_wire_strands' in post_fields and 'neutral_wire_strands' not in post_data:
                    post_data['neutral_wire_strands'] = get_val(['Strands (NW)', 'Strands (NW) '])

                # Spacing distances
                if 'spacing_d12' in post_fields and 'spacing_d12' not in post_data:
                    post_data['spacing_d12'] = sanitize_float(
                        get_val(['Spacing D12 (meters)', 'Spacing D12'])
                    )
                if 'spacing_d23' in post_fields and 'spacing_d23' not in post_data:
                    post_data['spacing_d23'] = sanitize_float(
                        get_val(['Spacing D23 (meters)', 'Spacing D23'])
                    )
                if 'spacing_d13' in post_fields and 'spacing_d13' not in post_data:
                    post_data['spacing_d13'] = sanitize_float(
                        get_val(['Spacing D13 (meters)', 'Spacing D13'])
                    )
                if 'spacing_d1n' in post_fields and 'spacing_d1n' not in post_data:
                    post_data['spacing_d1n'] = sanitize_float(
                        get_val(['Spacing D1n (meters)', 'Spacing D1n'])
                    )
                if 'spacing_d2n' in post_fields and 'spacing_d2n' not in post_data:
                    post_data['spacing_d2n'] = sanitize_float(
                        get_val(['Spacing D2n (meters)', 'Spacing D2n'])
                    )
                if 'spacing_d3n' in post_fields and 'spacing_d3n' not in post_data:
                    post_data['spacing_d3n'] = sanitize_float(
                        get_val(['Spacing D3n (meters)', 'Spacing D3n'])
                    )
                if 'spacing_dc1_c2' in post_fields and 'spacing_dc1_c2' not in post_data:
                    post_data['spacing_dc1_c2'] = sanitize_float(
                        get_val(['Spacing DC1-C2       (meters)', 'Spacing DC1-C2 (meters)', 'Spacing DC1-C2'])
                    )

                # Heights
                if 'height_h1' in post_fields and 'height_h1' not in post_data:
                    post_data['height_h1'] = sanitize_float(
                        get_val(['Height H1   (meters)', 'Height H1 (meters)', 'Height H1'])
                    )
                if 'height_h2' in post_fields and 'height_h2' not in post_data:
                    post_data['height_h2'] = sanitize_float(
                        get_val(['Height H2   (meters)', 'Height H2 (meters)', 'Height H2'])
                    )
                if 'height_h3' in post_fields and 'height_h3' not in post_data:
                    post_data['height_h3'] = sanitize_float(
                        get_val(['Height H3   (meters)', 'Height H3 (meters)', 'Height H3'])
                    )
                if 'height_hn' in post_fields and 'height_hn' not in post_data:
                    post_data['height_hn'] = sanitize_float(
                        get_val(['Height Hn   (meters)', 'Height Hn (meters)', 'Height Hn'])
                    )

                # Earth resistivity
                if 'earth_resistivity' in post_fields and 'earth_resistivity' not in post_data:
                    post_data['earth_resistivity'] = sanitize_float(
                        get_val(['Earth Resistivity (Ohm-meter)', 'Earth Resistivity'])
                    )

                # Required for a post: pole_number, lat, lng
                # 1) Try explicit pole_number
                pole_number = post_data.get('pole_number') or get_val(['Pole Number', 'pole number', 'Pole#'])
                # 2) If missing, treat From_Bus_ID as the pole_number (your data model)
                if not pole_number:
                    pole_number = get_val(['from_bus_id', 'From_Bus_ID', 'From Bus ID', 'from bus id'])
                # 3) Fallbacks: To_Bus_ID or Primary Bus ID if still empty
                if not pole_number:
                    pole_number = get_val(['to_bus_id', 'To_Bus_ID', 'To Bus ID', 'to bus id']) \
                                  or get_val(['Primary Bus ID', 'primary_bus_id'])

                # Coordinates: support lat/lng and latitude/longitude
                lat = sanitize_float(post_data.get('lat') or get_val(['lat', 'latitude']))
                lng = sanitize_float(post_data.get('lng') or get_val(['lng', 'longitude', 'long']))

                if not pole_number or lat is None or lng is None:
                    # Not enough info to create a post from this row
                    stats['skipped'] += 1
                    continue

                post_data['pole_number'] = pole_number
                post_data['lat'] = lat
                post_data['lng'] = lng
                # Ensure we have a readable name
                if not post_data.get('name'):
                    post_data['name'] = f"Post {pole_number}"
                # If primary_bus_id is empty, align it with the pole/bus identifier
                if not post_data.get('primary_bus_id'):
                    post_data['primary_bus_id'] = pole_number

                # Upsert Post
                post = Post.query.filter_by(pole_number=pole_number).first()
                if post:
                    for k, v in post_data.items():
                        setattr(post, k, v)
                    stats['updated'] += 1
                else:
                    post = Post(**post_data)
                    db.session.add(post)
                    stats['created'] += 1
                seen_posts.add(pole_number)

                # --- DISTRIBUTION LINE SEGMENT FIELDS ---
                line_data = {}
                for field in line_fields:
                    val = get_val([field, field.replace('_', ' '), field.upper(), field.lower()])
                    if val is not None:
                        # Try to convert to float for numeric fields
                        if 'length' in field or 'spacing' in field or 'height' in field or 'resistivity' in field or 'lat' in field or 'lng' in field:
                            val = sanitize_float(val)
                        line_data[field] = val
                # Required: segment_id, from_bus_id, to_bus_id
                segment_id = line_data.get('segment_id') or get_val(['Primary Distribution Line Segment ID', 'segment_id'])
                from_bus_id = line_data.get('from_bus_id') or get_val(['From_Bus_ID', 'from_bus_id'])
                to_bus_id = line_data.get('to_bus_id') or get_val(['To_Bus_ID', 'to_bus_id'])
                if segment_id and from_bus_id and to_bus_id:
                    line_data['segment_id'] = segment_id
                    line_data['from_bus_id'] = from_bus_id
                    line_data['to_bus_id'] = to_bus_id
                    # Upsert DistributionLineSegment
                    line = DistributionLineSegment.query.filter_by(segment_id=segment_id).first()
                    if line:
                        for k, v in line_data.items():
                            setattr(line, k, v)
                        stats['updated'] += 1
                    else:
                        line = DistributionLineSegment(**line_data)
                        db.session.add(line)
                        stats['created'] += 1
                    seen_lines.add(segment_id)
            except Exception as e:
                stats['errors'].append({'row': row_idx, 'error': str(e)})
                stats['skipped'] += 1

        try:
            db.session.commit()
        except IntegrityError as e:
            db.session.rollback()
            return jsonify({'error': f'Database error: {e}'}), 500

        stats['message'] = f"Import complete: {stats['created']} created/updated, {stats['skipped']} skipped, {len(stats['errors'])} errors."
        return jsonify(stats)


        # Process rows
        for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
            try:
                # Extract key data with flexible column matching
                pole_number = (find_column_value(row, ['Pole Number', 'pole_number', 'pole number', 'Pole#'], header_map) or '').strip()
                
                if not pole_number:
                    stats['skipped'] += 1
                    continue
                
                # Get coordinates
                lat = sanitize_float(find_column_value(row, ['Lat', 'Latitude', 'lat', 'LAT'], header_map))
                lng = sanitize_float(find_column_value(row, ['Long', 'Longitude', 'Lon', 'lng', 'LONG'], header_map))
                
                if lat is None or lng is None:
                    stats['errors'].append({'row': row_idx, 'error': 'Invalid or missing coordinates'})
                    stats['skipped'] += 1
                    continue
                
                # Extract all infrastructure data
                post_data = {
                    'pole_number': pole_number,
                    'name': f"Post {pole_number}",
                    'lat': lat,
                    'lng': lng,
                    'area': find_column_value(row, ['Lat, Long', 'Area', 'area'], header_map),
                    'feeder': find_column_value(row, ['Feeder', 'feeder', 'Feeder Name'], header_map),
                    'pri_structure': find_column_value(row, ['Pri. Structure', 'Primary Structure'], header_map),
                    'pri_conductor_size': find_column_value(row, ['Conductor Size', 'Pri Conductor Size'], header_map),
                    'neutral_wire': find_column_value(row, ['Neutral Wire', 'Neutral'], header_map),
                    'configuration': find_column_value(row, ['Configuration', 'Config'], header_map),
                    'phasing': find_column_value(row, ['Phasing', 'Phase'], header_map),
                    'primary_bus_id': find_column_value(row, ['Primary Bus ID', 'Pri Bus'], header_map),
                    'sec_structure': find_column_value(row, ['Sec. Structure', 'Secondary Structure'], header_map),
                    'sec_conductor_size': find_column_value(row, ['Sec Conductor Size'], header_map),
                    'sec_type': find_column_value(row, ['Tpye', 'Type', 'Sec Type'], header_map),
                    'conductor_type': find_column_value(row, ['Conductor Type', 'Material'], header_map),
                    'sec_bus_id': find_column_value(row, ['Sec. Bus ID', 'Secondary Bus'], header_map),
                    'kva_rating': sanitize_float(find_column_value(row, ['kVA Rating', 'kVA', 'Rating'], header_map)),
                    'common_sole': find_column_value(row, ['Common/Sole', 'Common Sole'], header_map),
                    'transformer_bus_id': find_column_value(row, ['Transformer Bus ID', 'Transformer Bus'], header_map),
                    'transformer_phasing': find_column_value(row, ['Transformer Phasing'], header_map),
                    'grounding_rod': find_column_value(row, ['Grounding Rod', 'Grounding'], header_map),
                    'circuit': find_column_value(row, ['Circuit', 'Circuit ID'], header_map),
                    'l2_conductor_size': find_column_value(row, ['L2', 'L2 Size'], header_map),
                    'l1_conductor_size': find_column_value(row, ['L1', 'L1 Size'], header_map),
                    'meter_id': find_column_value(row, ['kWhr Meter', 'Meter', 'Meter ID', 'Serial Number'], header_map),
                    'meter_brand': find_column_value(row, ['Brand', 'Manufacturer', 'Meter Brand'], header_map),
                    'status': 'active',
                }
                
                # Remove None values
                post_data = {k: v for k, v in post_data.items() if v is not None}
                
                # Upsert by pole_number
                existing_post = Post.query.filter_by(pole_number=pole_number).first()
                if existing_post:
                    # Update existing
                    for key, value in post_data.items():
                        setattr(existing_post, key, value)
                    stats['updated'] += 1
                else:
                    # Create new
                    new_post = Post(**post_data)
                    db.session.add(new_post)
                    stats['created'] += 1
                
                db.session.flush()
            
            except Exception as e:
                stats['errors'].append({'row': row_idx, 'error': str(e)})
                stats['skipped'] += 1
                db.session.rollback()
        
        db.session.commit()
        stats['message'] = f"Import complete: {stats['created']} created, {stats['updated']} updated, {stats['skipped']} skipped"
        
        # Auto-infer connections from imported posts if any were created or updated
        if stats['created'] > 0 or stats['updated'] > 0:
            try:
                infer_connections_from_posts()
                app.logger.info(f"Auto-inferred connections from {stats['created'] + stats['updated']} posts")
            except Exception as e:
                app.logger.warning(f"Connection inference failed: {e}")
                # Not fatal - posts were imported even if connections couldn't be inferred
        
        return jsonify(stats), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500


def infer_connections_from_posts():
    """
    Auto-infer line connections from post data based on feeder and circuit.
    Creates Primary_to_Primary connections for posts in same feeder/circuit.
    """
    try:
        # Get all posts grouped by feeder and circuit
        posts = Post.query.all()
        if not posts:
            return
        
        # Group posts by (feeder, circuit) 
        groups = {}
        for post in posts:
            if post.feeder and post.circuit and post.pole_number:
                key = (post.feeder, post.circuit)
                if key not in groups:
                    groups[key] = []
                groups[key].append(post)
        
        # For each group, create connections between consecutive posts
        connection_count = 0
        for (feeder, circuit), group_posts in groups.items():
            # Sort by pole_number to get order
            try:
                group_posts.sort(key=lambda p: int(''.join(c for c in str(p.pole_number) if c.isdigit())) if any(c.isdigit() for c in str(p.pole_number)) else 0)
            except:
                pass  # If sorting fails, keep original order
            
            # Create connections between sequential posts
            for i in range(len(group_posts) - 1):
                from_post = group_posts[i]
                to_post = group_posts[i + 1]
                
                from_bus = from_post.primary_bus_id or f"P{from_post.pole_number}"
                to_bus = to_post.primary_bus_id or f"P{to_post.pole_number}"
                
                # Check if connection already exists
                existing = LineConnection.query.filter_by(
                    from_bus=from_bus,
                    to_bus=to_bus,
                    connection_type='Primary_to_Primary'
                ).first()
                
                if not existing:
                    conn = LineConnection(
                        from_bus=from_bus,
                        to_bus=to_bus,
                        connection_type='Primary_to_Primary',
                        feeder=feeder,
                        circuit=circuit
                    )
                    db.session.add(conn)
                    connection_count += 1
        
        db.session.commit()
        app.logger.info(f"Inferred {connection_count} new connections from posts")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Connection inference error: {e}")
        raise


@app.route('/api/distribution-lines/bulk-import', methods=['POST'])
@admin_required
def api_distribution_lines_bulk_import():
    """
    Bulk import distribution line segment data from CSV or Excel file.
    Expected columns (case-insensitive): segment_id, from_bus_id, to_bus_id, phasing, configuration,
    system_grounding_type, length_meters, conductor_type, conductor_size, conductor_unit, conductor_strands,
    neutral_wire_type, neutral_wire_size, neutral_wire_unit, neutral_wire_strands,
    spacing_d12, spacing_d23, spacing_d13, spacing_d1n, spacing_d2n, spacing_d3n, spacing_dc1_c2,
    height_h1, height_h2, height_h3, height_hn, earth_resistivity
    Returns: { created: int, updated: int, skipped: int, errors: [{row, error}] }
    """
    from werkzeug.utils import secure_filename
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    filename = secure_filename(file.filename).lower()
    is_excel = filename.endswith(('.xlsx', '.xls'))
    is_csv = filename.endswith('.csv')
    
    if not (is_excel or is_csv):
        return jsonify({'error': 'File must be CSV or Excel (.xlsx, .xls)'}), 400
    
    try:
        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
        rows = []
        
        # Parse file
        if is_excel:
            from openpyxl import load_workbook
            workbook = load_workbook(file.stream)
            worksheet = workbook.active
            header = None
            header_map = {}
            
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if header is None:
                    # First row is header
                    header = tuple((str(v).strip().lower() if v else '') for v in row)
                    # Build mapping: column_name -> index
                    for idx, col_name in enumerate(header):
                        header_map[col_name] = idx
                    continue
                rows.append(row)
        else:  # CSV
            import csv as csv_module
            stream = io.StringIO(file.stream.read().decode('utf-8', errors='replace'))
            reader = csv_module.DictReader(stream)
            if not reader.fieldnames:
                return jsonify({'error': 'CSV file is empty or invalid'}), 400
            
            # Normalize field names to lowercase
            normalized_fields = {f.lower(): f for f in reader.fieldnames}
            header_map = {k: normalized_fields.get(k, k) for k in normalized_fields}
            rows = list(reader)
        
        # Validate required columns
        required_cols = ['segment_id', 'from_bus_id', 'to_bus_id']
        missing = [c for c in required_cols if c not in header_map]
        if missing:
            return jsonify({'error': f'CSV/Excel must include columns: {", ".join(required_cols)}'}), 400
        
        # Process rows
        for row_idx, row in enumerate(rows, start=2):
            try:
                # Extract values
                if isinstance(row, dict):
                    # CSV
                    segment_id = (row.get(header_map.get('segment_id', '')) or '').strip()
                    from_bus_id = (row.get(header_map.get('from_bus_id', '')) or '').strip()
                    to_bus_id = (row.get(header_map.get('to_bus_id', '')) or '').strip()
                else:
                    # Excel
                    segment_id = str(row[header_map.get('segment_id', 0)] or '').strip() if 'segment_id' in header_map else ''
                    from_bus_id = str(row[header_map.get('from_bus_id', 1)] or '').strip() if 'from_bus_id' in header_map else ''
                    to_bus_id = str(row[header_map.get('to_bus_id', 2)] or '').strip() if 'to_bus_id' in header_map else ''
                
                if not (segment_id and from_bus_id and to_bus_id):
                    stats['errors'].append({'row': row_idx, 'error': 'segment_id, from_bus_id, to_bus_id are required'})
                    stats['skipped'] += 1
                    continue
                
                # Extract optional fields with proper type conversion
                def get_float(field_name, row_data, header_map, row):
                    if isinstance(row_data, dict):
                        val = row_data.get(header_map.get(field_name, ''), '')
                    else:
                        idx = header_map.get(field_name)
                        val = row[idx] if idx is not None and idx < len(row) else None
                    if val is None or val == '':
                        return None
                    try:
                        return float(val)
                    except (TypeError, ValueError):
                        return None
                
                def get_string(field_name, row_data, header_map, row):
                    if isinstance(row_data, dict):
                        val = row_data.get(header_map.get(field_name, ''), '')
                    else:
                        idx = header_map.get(field_name)
                        val = row[idx] if idx is not None and idx < len(row) else None
                    return (str(val).strip() if val else '') or None
                
                # Build line segment object
                line_data = {
                    'segment_id': segment_id,
                    'from_bus_id': from_bus_id,
                    'to_bus_id': to_bus_id,
                    'phasing': get_string('phasing', row, header_map, row),
                    'configuration': get_string('configuration', row, header_map, row),
                    'system_grounding_type': get_string('system_grounding_type', row, header_map, row),
                    'length_meters': get_float('length_meters', row, header_map, row),
                    'conductor_type': get_string('conductor_type', row, header_map, row),
                    'conductor_size': get_string('conductor_size', row, header_map, row),
                    'conductor_unit': get_string('conductor_unit', row, header_map, row),
                    'conductor_strands': get_string('conductor_strands', row, header_map, row),
                    'neutral_wire_type': get_string('neutral_wire_type', row, header_map, row),
                    'neutral_wire_size': get_string('neutral_wire_size', row, header_map, row),
                    'neutral_wire_unit': get_string('neutral_wire_unit', row, header_map, row),
                    'neutral_wire_strands': get_string('neutral_wire_strands', row, header_map, row),
                    'spacing_d12': get_float('spacing_d12', row, header_map, row),
                    'spacing_d23': get_float('spacing_d23', row, header_map, row),
                    'spacing_d13': get_float('spacing_d13', row, header_map, row),
                    'spacing_d1n': get_float('spacing_d1n', row, header_map, row),
                    'spacing_d2n': get_float('spacing_d2n', row, header_map, row),
                    'spacing_d3n': get_float('spacing_d3n', row, header_map, row),
                    'spacing_dc1_c2': get_float('spacing_dc1_c2', row, header_map, row),
                    'height_h1': get_float('height_h1', row, header_map, row),
                    'height_h2': get_float('height_h2', row, header_map, row),
                    'height_h3': get_float('height_h3', row, header_map, row),
                    'height_hn': get_float('height_hn', row, header_map, row),
                    'earth_resistivity': get_float('earth_resistivity', row, header_map, row),
                }
                
                # Upsert by segment_id
                line = DistributionLineSegment.query.filter_by(segment_id=segment_id).first()
                if line:
                    # Update existing
                    for key, value in line_data.items():
                        if value is not None:
                            setattr(line, key, value)
                    stats['updated'] += 1
                else:
                    # Create new
                    line = DistributionLineSegment(**line_data)
                    db.session.add(line)
                    stats['created'] += 1
            
            except Exception as e:
                stats['errors'].append({'row': row_idx, 'error': str(e)})
                stats['skipped'] += 1
        
        db.session.commit()
        return jsonify(stats), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500


@app.route('/api/transformers/by-bus/<bus_id>', methods=['GET'])
def api_transformers_by_bus(bus_id):
    """
    Return all DistributionTransformer records linked to a given bus ID.

    A transformer is considered linked when its from_primary_bus_id matches the bus_id.
    Response JSON:
      {
        "bus_id": "...",
        "count": <int>,
        "transformers": [ { transformer fields ... }, ... ]
      }
    Always returns JSON (never HTML) so the frontend fetch().json() call is safe.
    """
    try:
        transformers = DistributionTransformer.query.filter_by(from_primary_bus_id=bus_id).all()
        return jsonify({
            'bus_id': bus_id,
            'count': len(transformers),
            'transformers': [t.to_dict() for t in transformers],
        }), 200
    except Exception as e:
        return jsonify({'error': str(e), 'bus_id': bus_id}), 500


@app.route('/api/distribution-lines', methods=['GET'])
def api_distribution_lines():
    """GET: list all distribution line segments"""
    try:
        lines = DistributionLineSegment.query.order_by(DistributionLineSegment.id.desc()).all()
        return jsonify([line.to_dict() for line in lines])
    except Exception as e:
        # Table may not exist if migration not run; return empty list with warning
        app.logger.warning("Distribution lines query failed (table may not exist): %s", e)
        return jsonify([]), 200


@app.route('/api/transformers/bulk-import', methods=['POST'])
@admin_required
def api_transformers_bulk_import():
    """Bulk import distribution transformer data (example2.csv style).
    Flexible header matching: accepts headers like 'Distribution Transformer ID',
    'From Primary Bus ID', 'To Secondary Bus ID', 'Primary Phasing', etc.
    Returns: { created, updated, skipped, errors }
    """
    from werkzeug.utils import secure_filename

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    filename = secure_filename(file.filename).lower()
    is_excel = filename.endswith(('.xlsx', '.xls'))
    is_csv = filename.endswith('.csv')

    if not (is_excel or is_csv):
        return jsonify({'error': 'File must be CSV or Excel (.xlsx, .xls)'}), 400

    try:
        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
        rows = []
        header_map = {}

        if is_excel:
            from openpyxl import load_workbook
            workbook = load_workbook(file.stream)
            worksheet = workbook.active
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # Build header map: normalize names
                    for idx, col_name in enumerate(row):
                        if col_name is None:
                            continue
                        header_map[str(col_name).strip()] = idx
                    continue
                rows.append(row)
        else:
            import csv as csv_module
            stream = io.StringIO(file.stream.read().decode('utf-8', errors='replace'))
            reader = csv_module.DictReader(stream)
            if not reader.fieldnames:
                return jsonify({'error': 'CSV file is empty or invalid'}), 400
            # Keep original fieldnames in header_map for flexible lookup
            header_map = {fn: fn for fn in reader.fieldnames}
            rows = list(reader)

        if not rows:
            return jsonify({'error': 'File has no data rows'}), 400

        # For each row, use find_column_value to flexibly extract fields
        for row_idx, row in enumerate(rows, start=2):
            try:
                transformer_id = find_column_value(row, ['Transformer ID', 'Distribution Transformer ID', 'transformer_id', 'transformer id'], header_map) or ''
                from_bus = find_column_value(row, ['From Primary Bus ID', 'From_Primary_Bus_ID', 'from_primary_bus_id', 'from primary bus id', 'from_bus_id', 'from bus id'], header_map) or ''
                to_bus = find_column_value(row, ['To Secondary Bus ID', 'To_Secondary_Bus_ID', 'to_secondary_bus_id', 'to secondary bus id', 'to_bus_id', 'to bus id'], header_map) or ''

                if not transformer_id or not from_bus:
                    stats['errors'].append({'row': row_idx, 'error': 'transformer_id and from_primary_bus_id are required'})
                    stats['skipped'] += 1
                    continue

                def get_str(field_names):
                    return find_column_value(row, field_names, header_map)

                def get_float(field_names):
                    v = find_column_value(row, field_names, header_map)
                    try:
                        return float(v) if v not in (None, '') else None
                    except Exception:
                        return None

                data = {
                    'transformer_id': transformer_id,
                    'from_primary_bus_id': from_bus,
                    'to_secondary_bus_id': get_str(['To Secondary Bus ID', 'to_secondary_bus_id', 'to bus id']),
                    'primary_phasing': get_str(['Primary Phasing', 'primary_phasing']),
                    'secondary_phasing': get_str(['Secondary Phasing', 'secondary_phasing']),
                    'installation_type': get_str(['Installation Type', 'installation_type']),
                    'no_dts_in_bank': (lambda x: int(x) if x and str(x).strip().isdigit() else None)(find_column_value(row, ['No. DTs in Bank', 'No DTs in Bank', 'no_dts_in_bank'], header_map)),
                    'connection': get_str(['Connection', 'connection']),
                    'kva_rating': get_float(['KVA Rating', 'KVA Rating', 'KVA Rating', 'kva_rating']),
                    'primary_voltage_kv': get_float(['Primary Voltage Rating(kV)', 'Primary Voltage Rating (kV)', 'primary_voltage_kv']),
                    'secondary_voltage_kv': get_float(['Secondary Voltage Rating (kV)', 'secondary_voltage_kv']),
                    'primary_tap_kv': get_float(['Primary Tap Voltage (kV)', 'primary_tap_kv']),
                    'secondary_tap_kv': get_float(['Secondary Tap Voltage (kV)', 'secondary_tap_kv']),
                    'pct_z': get_float(['%Z', 'pct_z']),
                    'xr_ratio': get_float(['X/R Ratio', 'xr_ratio']),
                    'no_load_loss_kw': get_float(['No-Load Loss (kW)', 'no_load_loss_kw']),
                    'exciting_current_pct': get_float(['Exciting Current (%)', 'exciting_current_pct']),
                }

                # Upsert by transformer_id + from_primary_bus_id
                existing = DistributionTransformer.query.filter_by(transformer_id=transformer_id, from_primary_bus_id=from_bus).first()
                if existing:
                    for k, v in data.items():
                        if v is not None:
                            setattr(existing, k, v)
                    stats['updated'] += 1
                else:
                    new = DistributionTransformer(**{k: v for k, v in data.items() if v is not None})
                    db.session.add(new)
                    stats['created'] += 1

            except Exception as e:
                stats['errors'].append({'row': row_idx, 'error': str(e)})
                stats['skipped'] += 1

        db.session.commit()
        return jsonify(stats), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500


# --- Authentication: login / logout / user management ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login for admins (username + password) and viewers (username + access_code).
    Supports form POST (redirect flow) and JSON API POST (returns JSON).
    """
    if request.method == 'GET':
        # If already authenticated, redirect to dashboard; otherwise show the login screen
        if get_current_user():
            return redirect(url_for('dashboard'))
        return render_template('login.html')

    # Support both JSON API POST and form POST.
    # IMPORTANT: use silent=True so form submits don't raise 415 Unsupported Media Type.
    data = request.get_json(silent=True) or request.form or {}
    username = (data.get('username') or '').strip()
    if not username:
        # For form submits, redirect back with message (simple flow); for API, return JSON error
        if request.form:
            return render_template('login.html', error='username required'), 400
        return jsonify({'error': 'username required'}), 400

    user = User.query.filter_by(username=username).first()
    if not user:
        if request.form:
            return render_template('login.html', error='Invalid credentials'), 401
        return jsonify({'error': 'invalid credentials'}), 401

    # Admins: password required
    if user.is_admin():
        pw = data.get('password')
        if not pw or not check_password_hash(user.password_hash or '', pw):
            if request.form:
                return render_template('login.html', error='Invalid credentials'), 401
            return jsonify({'error': 'invalid credentials'}), 401
    else:
        # Viewer: access code required and account enabled
        code = (data.get('access_code') or '').strip()
        if not code or not user.access_enabled or user.access_code != code:
            if request.form:
                return render_template('login.html', error='Invalid credentials or disabled account'), 401
            return jsonify({'error': 'invalid credentials or disabled account'}), 401

    # success: create login session
    session.clear()
    session['user_id'] = user.id
    session['role'] = user.role

    # After successful login, redirect to dashboard for form POST; for API calls return JSON
    if request.form:
        return redirect(url_for('dashboard'))
    return jsonify({'result': 'ok', 'id': user.id, 'username': user.username, 'role': user.role})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/auth/whoami')
def whoami():
    u = get_current_user()
    if not u: return jsonify({'authenticated': False}), 200
    return jsonify({'authenticated': True, 'user': u.public_dict()})

# --- Admin UI pages ---
@app.route('/admin/viewers')
@admin_required
def admin_viewers():
    """Admin-only page to create viewer accounts and display access codes (visualization/auth only)."""
    viewers = User.query.filter_by(role='viewer').order_by(User.id.desc()).all()
    return render_template('admin_viewers.html', active_page='admin_viewers', viewers=viewers)


@app.route('/resources')
@admin_required
def resources():
    """Admin-only resources page for exporting data."""
    return render_template('resources.html', active_page='resources')

@app.route('/distribution-lines')
@admin_required
def distribution_lines():
    """Admin-only distribution line segments management page."""
    return render_template('distribution_lines.html', active_page='distribution_lines')

# Admin-only user management API
@app.route('/api/users', methods=['GET', 'POST'])
@admin_required
def api_users():
    if request.method == 'GET':
        users = User.query.order_by(User.id.asc()).all()
        return jsonify([u.public_dict() for u in users])
    # POST: create viewer user
    # Admin supplies a person's name; system generates a username + access code.
    data = request.get_json(silent=True) or {}
    display_name = (data.get('display_name') or '').strip()
    username = (data.get('username') or '').strip()
    if not username and not display_name:
        return jsonify({'error': 'display_name (or username) required'}), 400

    # Generate a username from display_name if not provided
    if not username:
        import re
        base = re.sub(r'[^a-z0-9]+', '.', display_name.lower()).strip('.')
        base = base or 'viewer'
        username = base
        # Ensure uniqueness by appending -2, -3, ...
        n = 2
        while User.query.filter_by(username=username).first() is not None:
            username = f"{base}-{n}"
            n += 1
    existing = User.query.filter_by(username=username).first()
    if existing:
        return jsonify({'error': 'username already exists'}), 400
    code = secrets.token_urlsafe(8)
    user = User(username=username, role='viewer', access_code=code, access_enabled=True)
    db.session.add(user)
    db.session.commit()
    return jsonify({'id': user.id, 'username': user.username, 'access_code': code, 'access_enabled': user.access_enabled}), 201

@app.route('/setup/create-admin', methods=['POST'])
def setup_create_admin():
    """Convenience endpoint available only in debug mode to create an initial admin account.
    Accepts JSON: {username, password}. Only usable when no admins exist.
    """
    if not app.debug:
        return jsonify({'error': 'not available'}), 403
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    if not username or not password:
        return jsonify({'error': 'username and password required'}), 400
    existing_admin = User.query.filter_by(role='admin').first()
    if existing_admin:
        return jsonify({'error': 'admin already exists'}), 400
    pw_hash = generate_password_hash(password)
    u = User(username=username, role='admin', password_hash=pw_hash)
    db.session.add(u)
    db.session.commit()
    return jsonify({'id': u.id, 'username': u.username, 'role': u.role}), 201

@app.route('/api/users/<int:user_id>/regenerate_code', methods=['POST'])
@admin_required
def api_user_regen_code(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'user not found'}), 404
    code = secrets.token_urlsafe(8)
    user.access_code = code
    db.session.commit()
    return jsonify({'id': user.id, 'username': user.username, 'access_code': code})

@app.route('/api/users/<int:user_id>/enable', methods=['POST'])
@admin_required
def api_user_enable(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'user not found'}), 404
    data = request.get_json() or {}
    enable = bool(data.get('enable'))
    user.access_enabled = enable
    db.session.commit()
    return jsonify({'id': user.id, 'username': user.username, 'access_enabled': user.access_enabled})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_user_delete(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'user not found'}), 404
    if user.role == 'admin':
        return jsonify({'error': 'cannot delete admin user'}), 403
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({'id': user.id, 'deleted': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/posts/<int:post_id>/status', methods=['POST'])
@admin_required
def api_post_status(post_id):
    """Toggle or set the status of a post.
    POST body (optional): { "action": "toggle" } or { "status": "Active" }
    Returns: { id, status }
    """
    try:
        post = Post.query.get(post_id)
        if not post:
            return jsonify({'error': 'post not found'}), 404
        data = request.get_json() or {}
        action = data.get('action', 'toggle')
        if action == 'toggle':
            s = (post.status or '').strip().lower()
            post.status = 'Inactive' if s == 'active' else 'Active'
        else:
            new_status = data.get('status')
            if new_status is None:
                return jsonify({'error': 'status or action required'}), 400
            post.status = str(new_status)
        db.session.commit()
        return jsonify({'id': post.id, 'status': post.status})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import_latlong', methods=['POST'])
@admin_required
def api_import_latlong():
    """Import coordinates from `latlongdata` table into `posts` table (upsert by post_id).
    Call via: POST /api/import_latlong
    Returns JSON with import stats.
    """
    try:
        from import_latlong import import_from_latlong
        stats = import_from_latlong()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# New endpoint to return raw latlongdata normalized to post_id/lat/lng
from sqlalchemy import text

@app.route('/api/latlongdata')
def api_latlongdata():
    """Return normalized rows from `latlongdata` table as JSON: [{post_id, lat, lng}, ...]
    Column names are detected; if not present, the first three columns are used in order.
    """
    try:
        desc = db.session.execute(text('DESCRIBE latlongdata')).fetchall()
        col_names = [row[0] for row in desc]
        lower_cols = [c.lower() for c in col_names]
        expected = ['post_id', 'latitude', 'longitude']
        mapping = {}
        for name in expected:
            if name in lower_cols:
                mapping[name] = col_names[lower_cols.index(name)]

        if not all(k in mapping for k in expected):
            if len(col_names) < 3:
                return jsonify({'error': 'latlongdata must have at least 3 columns'}), 400
            mapping = {'post_id': col_names[0], 'latitude': col_names[1], 'longitude': col_names[2]}

        sql = text(f"SELECT `{mapping['post_id']}` AS post_id, `{mapping['latitude']}` AS latitude, `{mapping['longitude']}` AS longitude FROM latlongdata")
        rows = db.session.execute(sql).fetchall()

        out = []
        for r in rows:
            # Support different Row types: mapping by name (SQLAlchemy Row._mapping) or tuple-like
            if hasattr(r, '_mapping'):
                mapping = r._mapping
            else:
                # Fallback: map by positional columns (0=post_id,1=latitude,2=longitude)
                mapping = {
                    'post_id': r[0] if len(r) > 0 else None,
                    'latitude': r[1] if len(r) > 1 else None,
                    'longitude': r[2] if len(r) > 2 else None,
                }

            try:
                pid_raw = mapping.get('post_id')
                pid = int(pid_raw) if pid_raw is not None else None
            except Exception:
                pid = None
            try:
                lat_raw = mapping.get('latitude')
                lng_raw = mapping.get('longitude')
                lat = float(lat_raw) if lat_raw is not None else None
                lng = float(lng_raw) if lng_raw is not None else None
            except Exception:
                lat = None
                lng = None
            # Skip header-like or fully-empty rows (e.g., first row with column names)
            if pid is None and lat is None and lng is None:
                continue
            out.append({'post_id': pid, 'lat': lat, 'lng': lng})

        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# --- Feeder (bus-based) visualization: Bus–Post mapping + Excel bus-to-bus → draw lines only when both buses mapped ---
# Visualization only; no electrical logic. Skipped connections are logged and returned in API.


@app.route('/api/feeder/lines')
def api_feeder_lines():
    """
    Resolve engineering feeder data (Excel bus-to-bus) using Bus–Post mapping.
    Returns line segments to draw only when both buses map to posts with coordinates.
    Response: { lines: [...], skipped: [...] }. Returns 200 with empty lines/skipped on error so map still loads.
    """
    try:
        from feeder_data import resolve_feeder_lines
        result = resolve_feeder_lines(app)
        return jsonify(result)
    except Exception as e:
        # Return 200 with empty data so the map loads; include error for console/debugging
        app.logger.warning("Feeder lines failed: %s", e)
        return jsonify({'lines': [], 'skipped': [], 'error': str(e)}), 200


@app.route('/api/bus_post_mapping', methods=['GET', 'POST'])
def api_bus_post_mapping():
    """
    GET: list all bus_id -> post_id mappings.
    POST: add or replace a mapping. Body: { "bus_id": str, "post_id": int }.
    """
    if request.method == 'GET':
        try:
            rows = BusPostMapping.query.all()
            out = [{'id': r.id, 'bus_id': r.bus_id, 'post_id': r.post_id} for r in rows]
            return jsonify(out)
        except Exception as e:
            # Table may not exist if migration not run; return empty list so callers don't break
            app.logger.warning("bus_post_mapping GET failed: %s", e)
            return jsonify([]), 200

    # POST: add/update mapping (one at a time for minimal code)
    # Only admins may add or update mappings
    if not g.get('current_user') or not g.current_user.is_admin():
        return jsonify({'error': 'admin required'}), 403
    try:
        data = request.get_json() or {}
        bus_id = (data.get('bus_id') or '').strip()
        try:
            post_id = int(data.get('post_id'))
        except (TypeError, ValueError):
            return jsonify({'error': 'post_id must be an integer'}), 400
        if not bus_id:
            return jsonify({'error': 'bus_id is required'}), 400
        existing = BusPostMapping.query.filter_by(bus_id=bus_id).first()
        if existing:
            existing.post_id = post_id
            db.session.commit()
            return jsonify({'id': existing.id, 'bus_id': existing.bus_id, 'post_id': existing.post_id}), 200
        m = BusPostMapping(bus_id=bus_id, post_id=post_id)
        db.session.add(m)
        db.session.commit()
        return jsonify({'id': m.id, 'bus_id': m.bus_id, 'post_id': m.post_id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# --- Data Management: Delete all data and reset IDs ---

@app.route('/api/data/delete-all', methods=['POST'])
@admin_required
def api_delete_all_data():
    """
    Delete ALL data from posts, connections, network lines, and raw lat/long data.
    Also resets all auto-increment IDs to 1.
    WARNING: This is a destructive operation!
    """
    try:
        # Delete all data from all tables
        app.logger.info("Starting deletion of all data...")
        
        LineConnection.query.delete()
        DistributionLineSegment.query.delete()
        DistributionTransformer.query.delete()
        BusPostMapping.query.delete()
        Meter.query.delete()
        
        # Try to delete from LatLongData if it exists
        try:
            from models import LatLongData
            if LatLongData:
                LatLongData.query.delete()
        except:
            pass  # Table might not exist
        
        Post.query.delete()
        db.session.commit()
        
        app.logger.info("All data deleted successfully from database")
        
        # Reset auto-increment IDs based on database engine
        try:
            db_url = app.config.get('SQLALCHEMY_DATABASE_URI', '')
            app.logger.info(f"Database URL type: {db_url}")
            
            tables_to_reset = [
                'post',
                'line_connection',
                'distribution_line_segment',
                'distribution_transformer',
                'bus_post_mapping',
                'meter',
                'latlongdata',
            ]
            reset_info = {}
            
            if not db_url:
                app.logger.warning("No database URL found")
                reset_info = {table: 'skipped - no db_url' for table in tables_to_reset}
            elif 'mysql' in db_url or 'mariadb' in db_url:
                # MySQL/MariaDB: use ALTER TABLE to reset AUTO_INCREMENT
                app.logger.info("Using MySQL/MariaDB ID reset method")
                for table in tables_to_reset:
                    try:
                        # Use backticks for table names in case of reserved words
                        db.session.execute(db.text(f'ALTER TABLE `{table}` AUTO_INCREMENT = 1'))
                        reset_info[table] = 'reset'
                        app.logger.info(f"Reset {table} AUTO_INCREMENT")
                    except Exception as e:
                        # If ALTER fails, try TRUNCATE as fallback (harder reset)
                        try:
                            app.logger.info(f"ALTER failed for {table}, trying TRUNCATE...")
                            db.session.execute(db.text(f'TRUNCATE TABLE `{table}`'))
                            reset_info[table] = 'reset (truncate)'
                            app.logger.info(f"Truncated and reset {table}")
                        except Exception as e2:
                            reset_info[table] = f'error: {str(e2)}'
                            app.logger.warning(f"Failed to reset {table}: {e2}")
                db.session.commit()
            elif 'sqlite' in db_url:
                # SQLite: delete from sqlite_sequence table to reset AUTOINCREMENT
                app.logger.info("Using SQLite ID reset method")
                for table in tables_to_reset:
                    try:
                        db.session.execute(db.text(f"DELETE FROM sqlite_sequence WHERE name='{table}'"))
                        reset_info[table] = 'reset'
                        app.logger.info(f"Reset {table} sequence")
                    except Exception as e:
                        reset_info[table] = 'no sequence entry (ok)'
                        app.logger.debug(f"No sequence for {table}: {e}")
                db.session.commit()
            else:
                app.logger.warning(f"Unknown database type: {db_url}")
                reset_info = {table: 'unknown db type' for table in tables_to_reset}
        except Exception as reset_err:
            app.logger.warning(f"ID reset warning: {reset_err}")
            reset_info['_error'] = str(reset_err)
            # Not fatal - data is still deleted even if ID reset has issues
        
        app.logger.info("Delete all operation completed successfully")
        return jsonify({
            'result': 'success',
            'message': 'All data deleted and IDs reset to 1',
            'tables_deleted': ['post', 'line_connection', 'distribution_line_segment', 'bus_post_mapping', 'meter', 'latlongdata'],
            'id_reset_status': reset_info
        }), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Delete all data failed: {type(e).__name__}: {str(e)}", exc_info=True)
        return jsonify({
            'error': 'delete_failed',
            'message': f'Failed to delete all data: {type(e).__name__}: {str(e)}'
        }), 500


# --- Inferred Line Connections: Network topology from infrastructure data ---

@app.route('/api/line-connections', methods=['GET'])
def api_line_connections():
    """
    Get all inferred line connections for network visualization.
    
    Query parameters:
    - feeder: Filter by feeder ID (e.g., "F6")
    - type: Filter by connection type (e.g., "Primary_to_Primary")
    - from_bus: Filter by from_bus ID
    - to_bus: Filter by to_bus ID
    
    Returns: {connections: [...], total: int}
    """
    try:
        query = LineConnection.query
        
        # Apply filters
        feeder = request.args.get('feeder')
        conn_type = request.args.get('type')
        from_bus = request.args.get('from_bus')
        to_bus = request.args.get('to_bus')
        
        if feeder:
            query = query.filter_by(feeder=feeder)
        if conn_type:
            query = query.filter_by(connection_type=conn_type)
        if from_bus:
            query = query.filter_by(from_bus=from_bus)
        if to_bus:
            query = query.filter_by(to_bus=to_bus)
        
        connections = query.all()
        total = len(connections)
        
        return jsonify({
            'connections': [c.to_dict() for c in connections],
            'total': total
        })
    except Exception as e:
        app.logger.warning("line_connections GET failed: %s", e)
        return jsonify({'error': str(e), 'connections': [], 'total': 0}), 200


@app.route('/api/network-geometry')
def api_network_geometry():
    """
    Get network geometry (lines + nodes) directly from the DB posts,
    using only stored coordinates and the engineering rules in
    `network_geometry_db.get_network_geometry`.

    This does NOT require pre-populated LineConnection records; it infers
    primary segments and transformer/secondary links from the Post table.
    """
    try:
        from network_geometry_db import get_network_geometry

        result = get_network_geometry(app)
        # Shape the response for the frontend: expose lines/nodes/stats
        return jsonify({
            'lines': result.get('lines', []),
            'nodes': len(result.get('nodes', [])),
            'stats': result.get('stats', {}),
        }), 200
    except Exception as e:
        app.logger.error(f"network_geometry failed: {e}")
        return jsonify({
            'lines': [],
            'nodes': 0,
            'stats': {'nodes': 0, 'edges': 0, 'total_length_meters': 0, 'by_type': {}},
            'error': str(e)
        }), 200


@app.route('/api/line-connections/stats', methods=['GET'])
def api_line_connections_stats():
    """
    Get statistics about line connections.
    
    Returns: {
        total_connections: int,
        by_type: {connection_type: count, ...},
        by_feeder: {feeder: count, ...},
        unique_buses: int
    }
    """
    try:
        from sqlalchemy import func
        
        total = LineConnection.query.count()
        
        # Count by type
        by_type_query = db.session.query(
            LineConnection.connection_type,
            func.count(LineConnection.id)
        ).group_by(LineConnection.connection_type).all()
        by_type = {t[0]: t[1] for t in by_type_query}
        
        # Count by feeder
        by_feeder_query = db.session.query(
            LineConnection.feeder,
            func.count(LineConnection.id)
        ).group_by(LineConnection.feeder).all()
        by_feeder = {t[0]: t[1] for t in by_feeder_query if t[0]}
        
        # Count unique buses
        bus_query = db.session.query(
            LineConnection.from_bus.distinct(),
            LineConnection.to_bus.distinct()
        )
        unique_buses_set = set()
        for conn in LineConnection.query.all():
            unique_buses_set.add(conn.from_bus)
            unique_buses_set.add(conn.to_bus)
        
        return jsonify({
            'total_connections': total,
            'by_type': by_type,
            'by_feeder': by_feeder,
            'unique_buses': len(unique_buses_set)
        })
    except Exception as e:
        app.logger.warning("line_connections stats failed: %s", e)
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    debug = str(os.getenv('FLASK_DEBUG', '')).lower() in ('1', 'true', 'yes') or os.getenv('FLASK_ENV') == 'development'
    app.run(debug=debug, host='0.0.0.0', port=int(os.getenv('PORT', 5000)))
